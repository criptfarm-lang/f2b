"""F2B — обработчик сообщений канала «Мониторинг» в procurement.lots.

План 2026-05-22, Фаза 1.4 (закрыта 2026-05-26). Cron-задача в боте «Эф»:
каждые 30 минут 9-19 МСК читает unprocessed сообщения из market_intel_messages,
парсит через Anthropic SDK (text/photo через Claude haiku-4-5),
записывает в procurement.lots, помечает обработанным.

PDF на этом этапе SKIPPED (требует chunking из-за лимита output tokens на больших
прайс-листах типа Мореодор 300+ позиций). Vision-обработка PDF планируется в
следующей итерации; пока для PDF используется ручной запуск скилла update-market-intel.
"""

import base64
import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from claude_ai import get_client

logger = logging.getLogger(__name__)

ANTHROPIC_MODEL = os.getenv("MARKET_INTEL_MODEL", "claude-haiku-4-5-20251001")
# Отдельная модель для PDF — Sonnet 4.6 видит native PDF, понимает таблицы.
# Haiku обрабатывает text/photo (дешевле). См. план 2026-05-28-автопарсер-pdf-прайсов.md.
ANTHROPIC_PDF_MODEL = os.getenv("MARKET_INTEL_PDF_MODEL", "claude-sonnet-4-6")

# ─── ENUM-перечни (зеркало procurement_app/migrations/001+002.sql) ─────────
SPECIES_ENUM = [
    # лососёвые
    "лосось", "форель", "кета", "горбуша", "нерка", "кижуч", "чавыча",
    # белая рыба
    "треска", "пикша", "минтай", "судак", "сибас", "дорадо", "палтус",
    "зубатка", "масляная", "тилапия", "пангасиус",
    # прочая рыба
    "угорь", "тунец", "скумбрия", "сельдь",
    # морепродукты
    "креветка", "кальмар", "осьминог", "гребешок", "мидия", "устрица",
    "краб", "омар", "лангустин",
    # икра / прочее
    "икра", "прочее",
    # расширение из 002 (Moreodor)
    "осётр", "белуга", "стерлядь", "баррамунди", "барабулька", "ледяная",
    "нототения", "окунь", "эсколар", "кобия", "клыкач", "угольная",
    "хек", "хоки", "мойва", "навага", "корюшка", "омуль", "камбала",
    "сайда", "терпуг", "щука", "марлин", "рыба-меч", "тилапия-галилео",
]

REGION_ENUM = [
    "Чёрное море РФ", "Мурманск", "Карелия", "Северная Осетия",
    "Северо-Запад РФ", "Дальний Восток РФ", "Каспий",
    "Армения", "Беларусь", "Казахстан",
    "Иран", "Турция-озеро", "Турция-море", "Грузия", "Норвегия",
    "Фарерские о-ва", "Чили", "Аргентина", "Эквадор", "Перу",
    "Китай", "Вьетнам", "Таиланд", "Индия", "Индонезия", "Бангладеш",
    "прочее",
    # расширение из 002
    "Уругвай", "Корея", "Мавритания", "Египет", "РФ", "импорт",
    "Вьетнам/Китай",
]

PROCESSING_ENUM = [
    "НПСГ", "ПСГ", "ПБГ", "Б/Г",
    "Trim PR", "Trim A", "Trim B", "Trim C", "Trim D",
    "стейк", "кусок", "тушка", "хвост", "unspecified",
    # расширение из 002
    "НР", "HGT", "HLSO", "HOSO", "PDTO", "PDTL",
    "филе б/к", "филе н/к", "филе с/к", "кубик", "фарш",
    "лоин", "голова", "суповой набор", "J-CUT", "бабочка",
    "ломтики", "Saku",
]

STATE_ENUM = [
    "охл", "IQF", "блок", "глубокая-заморозка", "сушёный",
    "морская заморозка", "береговая заморозка", "полублок",
]

PRODUCT_FORM_ENUM = [
    "сырьё", "слабосоль", "сильносоль", "х/к", "г/к",
    "сухой-засол", "вяленое", "консервы",
    "жаренный", "с соусом",
]

# Маппинг подсказок поставщиков → slug в procurement.suppliers.
# Расширяется по мере появления новых поставщиков в канале.
# ВАЖНО: маппинг проверяется в порядке итерации, более специфичные ключи
# должны идти РАНЬШЕ общих (например "юнифрост иран" > "юнифрост").
SUPPLIER_HINT_MAP = {
    # Премиум лосось Мурманск
    "альфа-марин": "alfa-marin",
    "альфа марин": "alfa-marin",
    "alfa marin": "alfa-marin",
    "alfa-marin": "alfa-marin",
    # Мореодор
    "мореодор": "moreodor",
    "moreodor": "moreodor",
    # fish2o / Inarctica / Мурман Экспорт
    "мурман экспорт": "fish2o-murman-export",
    "fish2o": "fish2o-murman-export",
    "inarctica": "fish2o-murman-export",
    # Лаки Фиш
    "лаки фиш": "luckyfish",
    "luckyfish": "luckyfish",
    # Тем Групп
    "тем групп": "tem-grupp-artur",
    "tem grupp": "tem-grupp-artur",
    "артур": "tem-grupp-artur",
    # Карелия
    "федоренко": "fedorenko-karelia",
    "карелия федоренко": "fedorenko-karelia",
    # Смарт ЧМ
    "смарт": "smart-chm",
    "smart": "smart-chm",
    # Иран (несколько поставщиков)
    "юнифрост иран": "junifrost-iran",  # порядок важен: до общего "юнифрост"
    "junifrost iran": "junifrost-iran",
    "ирна": "irna",
    "эрам": "eram-iran",
    # Юнифрост Россия (другой поставщик, не путать с Iran)
    "unifrost.ru": "unifrost-ru",
    "unifrostru": "unifrost-ru",
    "григор": "unifrost-ru",
    "юнифрост": "unifrost-ru",  # дефолт без квалификатора — РФ
    "unifrost": "unifrost-ru",
    # Sky Fish
    "sky fish": "sky-fish",
    "skyfish": "sky-fish",
    # Фарватер (СПб, swfish.ru)
    "фарватер": "farvater",
    "farvater": "farvater",
    "swfish": "farvater",
    "swfish.ru": "farvater",
    "swfishru": "farvater",
    # DEFA Group (продаётся через Неву)
    "defa": "defa-fish",
    "дефа": "defa-fish",
    # Нева-Опт
    "нева опт": "нева-опт",
    "нева-опт": "нева-опт",
    "neva": "нева-опт",
    # Настоящая Рыбная Компания (Шикотан)
    "настоящая рыбная": "nrk",
    "шикотан": "nrk",
    "nrk": "nrk",
    # UltraFish
    "ultrafish": "ultrafish",
    "ультрафиш": "ultrafish",
    # Остров-Фиш
    "остров": "ostrov-fish",
    "ostrov": "ostrov-fish",
    # Прочие (точечные контакты)
    "хотенко": "khotenko-andrey",
    "волна": "volna-vladimir",
    "сити-ритейл": "sk-retail",
    "ск ритейл": "sk-retail",
    "дмитрий": "dmitriy-turkey-georgia",
}


def _build_system_prompt() -> str:
    return f"""Ты парсер прайс-листа поставщика рыбы / морепродуктов для F2B (АО ФИШ ТУ БИЗНЕС).

Из текста / фото / PDF извлеки СПИСОК лотов. Один лот = одна товарная позиция одной ступени цены.

Закрытые ENUM-перечни (используй ТОЛЬКО эти значения; если нет подходящего — `прочее`/`unspecified` с пометкой `confidence_self: needs-review`):

species: {", ".join(SPECIES_ENUM)}
regions: {", ".join(REGION_ENUM)}
processing: {", ".join(PROCESSING_ENUM)}
state: {", ".join(STATE_ENUM)}
product_form: {", ".join(PRODUCT_FORM_ENUM)}

ПОДСКАЗКИ для маппинга:
- HG / HGT (head gutted) → processing="ПБГ" или "HGT" если упоминается tail-off
- HON (head-on) → processing="ПСГ"
- HLSO/HOSO/PDTO/PDTL — для креветок, как processing
- ОХЛ / охлаждённое → state="охл"
- IQF / инд. заморозка / индивидуальная заморозка → state="IQF"
- block / блочная → state="блок"
- 21/25, 30/40 и т.п. для креветок — это weight_class (счёт штук на фунт), НЕ вес упаковки
- Вес тары («~30 кг», «8х1 кг», «12*1 кг») — НИКОГДА не уходит в weight_class. Только в conditions.
- «2-3 кг», «3,6-4,5», «1.5-2.0» — это weight_class рыбы
- «PREM» в названии — пометка премиум-сорта; включи в notes, не в species
- Регион «РФ» = Россия общая (если конкретного региона нет)
- «ДВ» = "Дальний Восток РФ"
- Сёмга / Семга = лосось (subspecies="атлантический")

ПРАВИЛА ЦЕНЫ (КРИТИЧНО — F2B работает в МСК):
1. Если в строке прайса несколько городов (СПб/МСК/Липецк/ДВ) → берём цену для МСК. В conditions запиши "МСК".
2. Если цена одна без указания города → берём её, в conditions локацию не указываем.
3. Если в строке несколько цен по условиям оплаты (предоплата/отсрочка) → берём ПРЕДОПЛАТУ как опт. В conditions запиши "предоплата".
4. Если в Caption (text_raw из TG) есть модификатор цены ("+7 руб для МСК", "к цене +7", "по Москве дороже на N", "акция -10%") — ПРИМЕНИ к цене для МСК и зафиксируй факт в notes (например "+7 ₽/кг для МСК из caption").
5. Если в прайсе несколько городов и для МСК ПУСТО (нет цены, нет ● маркера) — строку ПРОПУСКАЕМ (не записываем в lots).
6. ИГНОРИРУЕМ: вес тары/упаковки, период вылова, штрих-коды, артикулы (это либо в notes, либо вообще не записываем).

ПРАВИЛА GROUP-BY ЗАГОЛОВКОВ (Fix E, 29.05 — типичная структура Невы / СМАРТ ФИШ):
Многие прайсы используют структуру:
  «Креветка ваннамей ОЧИЩ:»  ← заголовок-группировщик, species + processing
    21/25 - 680₽
    31/40 - 540₽
    41/50 - 460₽
Каждая строка под заголовком — отдельный лот. Species/processing/state из заголовка
НАСЛЕДУЮТСЯ для всех строк группы пока не встретится новый заголовок другого species.
То же для секций «КРАСНАЯ РЫБА» / «БЕЛАЯ РЫБА» / «МОРЕПРОДУКТЫ» — это не лоты, это
контекст species. Извлекай ВСЕ строки группы как отдельные лоты с заголовочным
species, а не только первую.

ПРАВИЛО WEIGHT_CLASS (Fix C, 29.05):
weight_class — ОБЯЗАТЕЛЬНОЕ поле в БД, NOT NULL. Если калибр в прайсе не указан
вообще (например только species + цена) — поставь "unspecified". Если в прайсе
штучный счёт креветок (21/25, 30/40) — это и есть weight_class ("21/25" как строка),
а НЕ null.

Формат ответа — СТРОГО JSON-объект, БЕЗ markdown-fences, БЕЗ переносов внутри объектов lot (один lot на строку, compact JSON для экономии токенов):

{{
  "supplier_hint": "название поставщика как видишь",
  "received_at": "YYYY-MM-DD или null",
  "lots": [
    {{"species":"из enum","weight_class":"вес/калибр","processing":"из enum","state":"из enum","price_rub_kg":число,"raw_text":"исходный текст"}}
  ]
}}

ОБЯЗАТЕЛЬНЫЕ ПОЛЯ в каждом lot: species, weight_class, processing, state, price_rub_kg, raw_text.

ОПЦИОНАЛЬНЫЕ (добавляй только если есть значение, иначе НЕ включай ключ): subspecies, region, product_form (default «сырьё»), volume_tier, conditions, confidence_self (default «confirmed»; ставь «needs-review» при сомнении), notes.

Если не прайс — верни {{"supplier_hint":null,"received_at":null,"lots":[]}}.
"""


_HINT_RE = __import__("re").compile(r'"supplier_hint"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)"')
_RECV_RE = __import__("re").compile(r'"received_at"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)"')


def _parse_truncated_json(text: str) -> dict:
    """Восстанавливает максимум полных lot-объектов из обрезанного JSON-ответа.

    Идея: ответ имеет вид {"supplier_hint":...,"received_at":...,"lots":[{...},{...},{...
    где последний lot оборван. Ищем подстроку "lots":[ — всё ДО неё уходит в head;
    после неё пытаемся последовательно дожевать lots, пока json.loads не упадёт.

    Fix A (29.05): supplier_hint и received_at извлекаем regex'ом, а не json.loads
    головки. Раньше при наличии спецсимволов внутри hint json.loads падал и hint=null
    → SkyFish-style fail: лоты есть, но slug не определён, весь batch выкидывался.
    """
    fallback = {"supplier_hint": None, "received_at": None, "lots": []}
    marker = '"lots"'
    pos = text.find(marker)
    if pos < 0:
        return fallback
    # найти открывающую [
    br = text.find("[", pos)
    if br < 0:
        return fallback

    head_text = text[:br + 1]
    # Regex по head — устойчиво к truncation внутри hint и спецсимволам.
    hint_match = _HINT_RE.search(head_text)
    recv_match = _RECV_RE.search(head_text)
    supplier_hint = hint_match.group(1) if hint_match else None
    received_at = recv_match.group(1) if recv_match else None

    lots = []
    i = br + 1
    n = len(text)
    while i < n:
        # пропускаем пробелы и запятые
        while i < n and text[i] in " ,\n\r\t":
            i += 1
        if i >= n or text[i] != "{":
            break
        # ищем парную закрывающую } учитывая строки/escape
        depth = 0
        in_str = False
        esc = False
        j = i
        while j < n:
            ch = text[j]
            if in_str:
                if esc:
                    esc = False
                elif ch == "\\":
                    esc = True
                elif ch == '"':
                    in_str = False
            else:
                if ch == '"':
                    in_str = True
                elif ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        j += 1
                        break
            j += 1
        if depth != 0:
            # объект оборван
            break
        chunk = text[i:j]
        try:
            lots.append(json.loads(chunk))
        except json.JSONDecodeError:
            break
        i = j

    return {"supplier_hint": supplier_hint, "received_at": received_at, "lots": lots}


async def _call_claude(content, max_tokens: int = 8192, model: Optional[str] = None,
                       _debug_label: str = "", _debug_db=None) -> dict:
    """Универсальный вызов Anthropic API. content — список content-blocks.
    model — если None, используется ANTHROPIC_MODEL (Haiku по умолчанию).
    _debug_label + _debug_db — если заданы и lots=[], raw-ответ модели пишется в bot_settings."""
    client = get_client()
    used_model = model or ANTHROPIC_MODEL
    response = await client.messages.create(
        model=used_model,
        max_tokens=max_tokens,
        system=_build_system_prompt(),
        messages=[{"role": "user", "content": content}],
    )
    text = response.content[0].text
    raw_text = text  # сохраняем до strip для диагностики
    # Иногда Claude добавляет ```json ... ``` обёртки — снимаем
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text
        text = text.rsplit("```", 1)[0].strip()
        if text.startswith("json"):
            text = text[4:].lstrip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as e:
        # Robust-fallback: ответ обрезан max_tokens'ом посередине lot'а.
        # Ищем последний полный объект в lots: и закрываем массив + объект.
        parsed = _parse_truncated_json(text)
        if parsed.get("lots"):
            logger.warning(
                f"market_intel: JSON truncated, recovered {len(parsed['lots'])} lots from prefix"
            )
        else:
            logger.error(f"market_intel: JSON decode failed: {e}\ntext={text[:500]}")

    # Debug: если lots пустой и есть метка + db — кладём raw-ответ в bot_settings
    # (Amvera логи труднодоступны локально, БД — наш единственный канал).
    if _debug_label and _debug_db is not None and not parsed.get("lots"):
        try:
            usage = response.usage
            snippet = raw_text[:1500].replace("\n", "\\n")
            debug_value = (
                f"model={used_model} "
                f"in={usage.input_tokens} out={usage.output_tokens} "
                f"supplier_hint={parsed.get('supplier_hint')!r} "
                f"raw={snippet}"
            )
            _debug_db._execute(
                "INSERT INTO bot_settings(key, value) VALUES (%s, %s) "
                "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                (f"market_intel_debug_{_debug_label}", debug_value[:3000]),
            )
        except Exception as dbg_err:
            logger.warning(f"market_intel debug write failed: {dbg_err!r}")
    return parsed


async def parse_text_message(text: str) -> dict:
    return await _call_claude([
        {"type": "text", "text": f"Текст сообщения из канала «Мониторинг»:\n\n{text}"}
    ])


async def parse_photo_message(image_path: str, caption: str = "") -> dict:
    img_b64 = base64.b64encode(Path(image_path).read_bytes()).decode()
    media_type = "image/jpeg" if image_path.lower().endswith((".jpg", ".jpeg")) else "image/png"
    return await _call_claude([
        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
        {"type": "text", "text": f"Фото прайс-листа из канала «Мониторинг».\nCaption (если есть): {caption or '(нет)'}\nИзвлеки все товарные позиции."},
    ])


def _flatten_xls_to_text(xls_path: str, max_chars: int = 80000) -> str:
    """Открывает XLS/XLSX и возвращает плоский text-дамп всех листов/строк/ячеек.

    Sonnet не умеет читать XLS как native document, поэтому мы делаем flatten:
    ```
    === Sheet: 'Прайс' ===
    R1: ФАРВАТЕР | ООО Фарватер | ИНН 7813266850 | Санкт-Петербург
    R2:  | (812) 500-05-65 | swfish.ru
    ...
    R6: # | МОРЕПРОДУКТЫ | ПРОИЗВОДИТЕЛЬ | НАЛИЧИЕ | СПб | Мск
    R7: 1 | Капуста МОРСКАЯ 16+ | Китай | склад | 505,00 | 509,00
    ```
    Sonnet по контексту понимает шапку (R6 — header) и cell-positions.
    Обрезаем длину чтобы не превысить input window Sonnet (~200k токенов).
    """
    ext = xls_path.lower().rsplit(".", 1)[-1]
    parts = []
    if ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(xls_path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            parts.append(f"\n=== Sheet: {sheet.title!r} ===\n")
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                cells = [str(c).strip().replace("|", "/").replace("\n", " ") if c is not None else "" for c in row]
                line = f"R{r_idx}: " + " | ".join(cells)
                parts.append(line[:500])  # ограничиваем length одной строки
                if sum(len(p) for p in parts) > max_chars:
                    parts.append(f"\n... (truncated at {max_chars} chars)\n")
                    return "\n".join(parts)
        wb.close()
    elif ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(xls_path)
        for sheet in wb.sheets():
            parts.append(f"\n=== Sheet: {sheet.name!r} ===\n")
            for r_idx in range(sheet.nrows):
                row = sheet.row_values(r_idx)
                if all(str(c).strip() == "" for c in row):
                    continue
                cells = [str(c).strip().replace("|", "/").replace("\n", " ") for c in row]
                line = f"R{r_idx+1}: " + " | ".join(cells)
                parts.append(line[:500])
                if sum(len(p) for p in parts) > max_chars:
                    parts.append(f"\n... (truncated at {max_chars} chars)\n")
                    return "\n".join(parts)
    else:
        raise ValueError(f"unsupported xls extension: {ext}")
    return "\n".join(parts)


async def parse_xls_message(xls_path: str, caption: str = "", debug_label: str = "", debug_db=None) -> dict:
    """XLS/XLSX → flatten в text → Sonnet 4.6.

    Sonnet не имеет native XLS-блока (как PDF native document). Поэтому:
    1. openpyxl/xlrd читает все листы/строки/ячейки.
    2. _flatten_xls_to_text формирует текстовое представление с координатами cells.
    3. Sonnet получает это как user text + caption + system prompt.

    Шапка таблицы (как «# | МОРЕПРОДУКТЫ | ПРОИЗВОДИТЕЛЬ | НАЛИЧИЕ | СПб | Мск»)
    и merged cells читаются как plain row — Sonnet по контексту понимает что
    колонка «Мск» это цена для МСК (правила цены в system-prompt применяются).
    """
    try:
        flat_text = _flatten_xls_to_text(xls_path)
    except Exception as e:
        logger.error(f"market_intel: XLS flatten failed for {xls_path}: {e!r}")
        return {"supplier_hint": None, "received_at": None, "lots": []}
    if not flat_text.strip():
        logger.warning(f"market_intel: XLS {xls_path} пустой после flatten")
        return {"supplier_hint": None, "received_at": None, "lots": []}

    content = [
        {
            "type": "text",
            "text": (
                f"XLS/XLSX-прайс поставщика рыбы (плоский дамп всех листов).\n"
                f"Формат: «Rн: cell1 | cell2 | cell3 | ...» где Rн - номер строки.\n"
                f"Шапка таблицы (например «# | НАИМЕНОВАНИЕ | ПРОИЗВОДИТЕЛЬ | СПб | Мск») "
                f"даёт контекст что значит каждая колонка. Применяй правила цены из system "
                f"prompt: если колонки СПб и Мск разные — бери Мск; если города нет — бери "
                f"единственную цену. Игнорируй секции-заголовки и реквизиты, бери только товарные строки.\n\n"
                f"Caption от пересылающего: {caption or '(нет caption)'}\n\n"
                f"supplier_hint = название поставщика-производителя/дистрибьютора из шапки таблицы "
                f"(например «ФАРВАТЕР», «Юнифрост», «Sky Fish»). НЕ возвращай как supplier_hint "
                f"общие термины («прайс», «склад», «прайс-лист»), наши каналы («Мониторинг») "
                f"или наш юрлица («ФИШ ТУ БИЗНЕС»).\n\n"
                f"Дамп таблицы:\n{flat_text}"
            ),
        },
    ]
    return await _call_claude(content, max_tokens=16384, model=ANTHROPIC_PDF_MODEL,
                              _debug_label=debug_label, _debug_db=debug_db)


async def parse_pdf_message(pdf_path: str, caption: str = "", debug_label: str = "", debug_db=None) -> dict:
    """PDF → native document block в Claude Sonnet 4.6.

    План 2026-05-28: одна ступень — Sonnet видит PDF + caption (text_raw из TG)
    одним вызовом, разбирает шапки таблиц, объединённые ячейки, многоколоночность,
    применяет модификаторы цены из caption ("+7 для МСК"). Стоимость ~5 ₽/PDF.
    Предыдущая итерация (pdfplumber → плоский текст → Haiku) на табличных прайсах
    типа Мореодор давала lots=[] — Haiku не справлялся с грязным текстовым месивом.
    """
    pdf_bytes = Path(pdf_path).read_bytes()
    if not pdf_bytes:
        logger.warning(f"market_intel: PDF {pdf_path} пустой")
        return {"supplier_hint": None, "received_at": None, "lots": []}
    pdf_b64 = base64.b64encode(pdf_bytes).decode()

    content = [
        {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": pdf_b64,
            },
        },
        {
            "type": "text",
            "text": (
                f"PDF-прайс поставщика рыбы.\n"
                f"Caption от пересылающего (может содержать модификатор цены типа '+7 для МСК', "
                f"локацию, акции): {caption or '(нет caption)'}\n\n"
                f"Извлеки ВСЕ товарные позиции прайса. Применяй правила цены МСК / предоплата / "
                f"caption-модификаторы из system prompt.\n"
                f"supplier_hint = название КОНКРЕТНОГО поставщика-производителя или дистрибьютора "
                f"(берём из шапки PDF: «ООО Мореодор», «Sky Fish», «UltraFish», «DEFA Group» и т.п.). "
                f"НЕ возвращай как supplier_hint названия наших каналов («Мониторинг», «Эф»), "
                f"наш юрлица («ФИШ ТУ БИЗНЕС») или общие термины («прайс», «прайс-лист»). "
                f"Если в шапке нет явного бренда поставщика — supplier_hint=null.\n"
                f"Шапку поставщика, реквизиты, секции-заголовки CAPS — игнорируй "
                f"(используй как контекст species, но не записывай в lots)."
            ),
        },
    ]
    return await _call_claude(content, max_tokens=16384, model=ANTHROPIC_PDF_MODEL,
                              _debug_label=debug_label, _debug_db=debug_db)


def _supplier_slug_from_hint(hint: str) -> tuple[Optional[str], bool]:
    """Возвращает (slug, is_known). is_known=True если матч с SUPPLIER_HINT_MAP,
    False если slug сгенерирован из hint (новый/неопознанный поставщик).
    Лоты от is_known=False помечаются confidence=needs-review."""
    if not hint:
        return None, False
    h = hint.lower().strip()
    for needle, slug in SUPPLIER_HINT_MAP.items():
        if needle in h:
            return slug, True
    # Hint не в карте — генерим slug. Берём первое осмысленное слово, чистим.
    import re as _re
    # Игнорируем мусорные фразы из канала-обёртки.
    GARBAGE = (
        "мониторинг", "ооо фиш ту бизнес", "ооо «фиш ту бизнес»",
        "важно", "не забывайте", "что-то от поставщика",
        "кристина, сюда", "сюда пересылаем", "пересылаем",
    )
    for g in GARBAGE:
        if g in h:
            return None, False
    # Берём первые 30 символов, оставляем только буквы/цифры/пробелы/дефисы.
    cleaned = _re.sub(r"[^a-zа-я0-9\s\-]", "", h[:30]).strip()
    cleaned = _re.sub(r"\s+", "-", cleaned)
    if len(cleaned) < 2:
        return None, False
    return cleaned, False


def _insert_lots(db, lots: list[dict], supplier_id: str, msg_id: int,
                 source_file: str, received_at: str) -> tuple[int, int]:
    """INSERT в procurement.lots. Возвращает (inserted, skipped)."""
    try:
        valid_until = (datetime.strptime(received_at, "%Y-%m-%d").date()
                       + timedelta(days=14)).isoformat()
    except (ValueError, TypeError):
        # fallback: сегодня + 14 дней
        valid_until = (datetime.now().date() + timedelta(days=14)).isoformat()
        received_at = datetime.now().date().isoformat()

    db._ensure_connection() if hasattr(db, "_ensure_connection") else None
    conn = db.conn
    inserted, skipped = 0, 0
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO procurement.suppliers (slug, name) VALUES (%s, %s) ON CONFLICT (slug) DO NOTHING",
                (supplier_id, supplier_id),
            )
            for lot in lots:
                # валидация ENUM на стороне Python: если значение не в списке — needs-review
                species = lot.get("species")
                if species not in SPECIES_ENUM:
                    species, lot["confidence_self"] = "прочее", "needs-review"
                region = lot.get("region")
                if region and region not in REGION_ENUM:
                    region = None
                processing = lot.get("processing")
                if processing not in PROCESSING_ENUM:
                    processing = "unspecified"
                state = lot.get("state")
                if state not in STATE_ENUM:
                    state = "глубокая-заморозка"
                product_form = lot.get("product_form", "сырьё")
                if product_form not in PRODUCT_FORM_ENUM:
                    product_form = "сырьё"
                confidence = lot.get("confidence_self", "confirmed")
                if confidence not in ("confirmed", "pending", "needs-review", "archived"):
                    confidence = "confirmed"
                # Лоты вне нашей таксономии (готовая продукция, упаковка, прочее)
                # сразу в archived — не загружаем менеджера, но не теряем для /search
                # по raw_text если завтра прилетит запрос.
                if species == "прочее" or processing == "unspecified":
                    confidence = "archived"
                try:
                    price = float(lot["price_rub_kg"])
                except (KeyError, TypeError, ValueError):
                    skipped += 1
                    continue
                # Fix C (29.05): None в обязательных полях. dict.get("k", default)
                # возвращает None если ключ есть с None — нужен `or default`.
                # Креветка 21/25 (счёт штук) приходила с weight_class=None и роняла
                # весь batch через NotNullViolation + InFailedSqlTransaction.
                weight_class = lot.get("weight_class") or "unspecified"
                # Fix B (29.05): SAVEPOINT для каждого INSERT — иначе первая ошибка
                # абортит транзакцию и все последующие insert'ы валятся с
                # InFailedSqlTransaction (msg_24 СМАРТ ФИШ потерял 13 валидных лотов).
                cur.execute("SAVEPOINT lot_sp")
                try:
                    cur.execute(
                        """INSERT INTO procurement.lots (
                            species, subspecies, region, weight_class, processing, state,
                            product_form, price_rub_kg, volume_tier, conditions, supplier_id,
                            received_at, valid_until, msg_id, source_file, raw_text,
                            confidence, notes
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (supplier_id, COALESCE(msg_id, 0), species,
                                     COALESCE(region, 'прочее'::procurement.region_enum),
                                     weight_class, processing, state, price_rub_kg,
                                     COALESCE(volume_tier, ''), COALESCE(conditions, ''))
                        WHERE superseded_by_lot_id IS NULL DO NOTHING""",
                        (
                            species, lot.get("subspecies"), region,
                            weight_class,
                            processing, state, product_form, price,
                            lot.get("volume_tier"), lot.get("conditions"),
                            supplier_id, received_at, valid_until, msg_id,
                            source_file, lot.get("raw_text"),
                            confidence, lot.get("notes"),
                        ),
                    )
                    if cur.rowcount > 0:
                        inserted += 1
                    else:
                        skipped += 1
                    cur.execute("RELEASE SAVEPOINT lot_sp")
                except Exception as e:
                    cur.execute("ROLLBACK TO SAVEPOINT lot_sp")
                    logger.error(f"market_intel: insert lot failed: {e!r}; lot={lot}")
                    skipped += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return inserted, skipped


async def process_pending(db, limit: int = 20) -> dict:
    """Главная функция — обработать unprocessed сообщения."""
    pending = db.get_unprocessed_market_intel(limit=limit)
    stats = {"checked": len(pending), "processed": 0, "failed": 0,
             "skipped_pdf": 0, "inserted_total": 0}
    for msg in pending:
        msg_id = msg["id"]
        mt = msg.get("msg_type")
        caption = msg.get("text_raw") or ""
        forward_from = msg.get("forward_from") or ""
        file_path = msg.get("file_path")
        try:
            if mt == "text":
                if not caption.strip():
                    logger.info(f"market_intel: msg {msg_id} empty text, skip")
                    db.mark_market_intel_processed(msg_id)
                    stats["processed"] += 1
                    continue
                result = await parse_text_message(caption)
            elif mt == "photo" and file_path and os.path.exists(file_path):
                result = await parse_photo_message(file_path, caption)
            elif mt == "document" and msg.get("file_ext") == "pdf":
                if not file_path or not os.path.exists(file_path):
                    logger.warning(f"market_intel: msg {msg_id} PDF file_path не найден ({file_path}), помечаю processed")
                    db.mark_market_intel_processed(msg_id)
                    stats["failed"] += 1
                    continue
                try:
                    result = await parse_pdf_message(file_path, caption,
                                                     debug_label=f"msg{msg_id}", debug_db=db)
                except Exception as e:
                    logger.exception(f"market_intel: msg {msg_id} PDF parse failed: {e!r}")
                    # НЕ помечаем processed — попробуем на следующем тике (вдруг временный сбой).
                    stats["failed"] += 1
                    continue
            elif mt == "document" and msg.get("file_ext") in ("xls", "xlsx"):
                if not file_path or not os.path.exists(file_path):
                    logger.warning(f"market_intel: msg {msg_id} XLS file_path не найден ({file_path}), помечаю processed")
                    db.mark_market_intel_processed(msg_id)
                    stats["failed"] += 1
                    continue
                try:
                    result = await parse_xls_message(file_path, caption,
                                                     debug_label=f"msg{msg_id}", debug_db=db)
                except Exception as e:
                    logger.exception(f"market_intel: msg {msg_id} XLS parse failed: {e!r}")
                    stats["failed"] += 1
                    continue
            else:
                logger.warning(f"market_intel: msg {msg_id} type={mt} unknown, skip")
                db.mark_market_intel_processed(msg_id)
                stats["processed"] += 1
                continue

            # Маппинг supplier.
            # Fix (29.05): forward_from приоритетнее чем supplier_hint от Sonnet.
            # Причина: forward_from = title исходного канала откуда переслали (например
            # "Мореодор", "fish2o X Inarctica"), это детерминированный сигнал из Telegram.
            # supplier_hint от Sonnet — эвристика из шапки PDF, может быть "Мониторинг"
            # (название нашего канала, попавшее в контекст PDF) или другим шумом.
            hint = forward_from or result.get("supplier_hint") or caption[:50]
            slug, is_known = _supplier_slug_from_hint(hint)
            # Fix XLS (29.05 вечер): если Sonnet распознал лоты но supplier_hint пустой —
            # сохраняем лоты под auto-slug `unknown-msgN` с пометкой needs-review,
            # чтобы менеджер мог принять/переименовать через /needs_review.
            # Раньше такой батч (msg=22 = 139 креветок) выбрасывался полностью.
            if not slug and result.get("lots"):
                slug = f"unknown-msg{msg_id}"
                is_known = False
                logger.warning(
                    f"market_intel: msg {msg_id} — slug пустой, но lots={len(result['lots'])}; "
                    f"сохраняю под auto-slug={slug}"
                )
            if not slug:
                logger.warning(f"market_intel: msg {msg_id} — slug пустой (hint={hint!r}); пропускаю needs-review")
                # Fix D (29.05): debug-write при пустом slug — иначе случай
                # «Sonnet распознал N лотов, но supplier_hint оборван → batch выкинут»
                # не оставляет следов в bot_settings.market_intel_debug_*.
                try:
                    sample = (result.get("lots") or [])[:3]
                    debug_value = (
                        f"reason=empty_slug hint={hint!r} forward_from={forward_from!r} "
                        f"file={file_path or 'msg_'+str(msg_id)} "
                        f"lots_parsed={len(result.get('lots') or [])} "
                        f"sample_lots={sample!r}"
                    )
                    db._execute(
                        "INSERT INTO bot_settings(key, value) VALUES (%s, %s) "
                        "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                        (f"market_intel_debug_msg{msg_id}", debug_value[:3000]),
                    )
                except Exception as dbg_err:
                    logger.warning(f"market_intel debug-empty-slug write failed: {dbg_err!r}")
                db.mark_market_intel_processed(msg_id)
                stats["failed"] += 1
                continue

            received_at = result.get("received_at")
            if not received_at:
                posted_at = msg.get("posted_at")
                received_at = posted_at.strftime("%Y-%m-%d") if posted_at else datetime.now().strftime("%Y-%m-%d")

            # Авто-slug (не из known map) → лоты помечаются needs-review.
            lots_list = result.get("lots", [])
            if not is_known:
                for lot in lots_list:
                    lot["confidence_self"] = "needs-review"
                    lot.setdefault("notes", "")
                    lot["notes"] = (lot["notes"] + f" [auto-slug from hint='{hint[:40]}']").strip()

            ins, skp = _insert_lots(db, lots_list, slug,
                                     msg_id, file_path or f"msg_{msg_id}", received_at)
            db.mark_market_intel_processed(msg_id)
            stats["processed"] += 1
            stats["inserted_total"] += ins
            logger.info(f"market_intel: msg {msg_id} → supplier={slug} (known={is_known}) lots_inserted={ins} skipped={skp}")
        except Exception as e:
            logger.exception(f"market_intel: msg {msg_id} failed: {e!r}")
            stats["failed"] += 1
    return stats


def _record_tick(db, status: str, detail: str = ""):
    """Heartbeat в bot_settings. Чтобы видеть из БД, что cron живой и когда
    он последний раз стартовал/завершился (без доступа к Amvera-логам)."""
    try:
        from zoneinfo import ZoneInfo
        now_iso = datetime.now(ZoneInfo("Europe/Moscow")).isoformat()
        db._execute(
            "INSERT INTO bot_settings(key, value) VALUES (%s, %s) "
            "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            (f"market_intel_{status}_at", f"{now_iso} {detail}".strip()),
        )
    except Exception as e:
        logger.warning(f"_record_tick: {e}")


async def market_intel_cron_job(app, db):
    """APScheduler entry-point. Вызывается каждые 30 мин 24/7."""
    logger.info("market_intel cron tick started")
    _record_tick(db, "tick_started")
    try:
        stats = await process_pending(db)
        logger.info(f"market_intel cron: {stats}")
        _record_tick(db, "tick_done", detail=str(stats))
    except Exception as e:
        logger.exception("market_intel cron crashed")
        _record_tick(db, "tick_crashed", detail=str(e))
